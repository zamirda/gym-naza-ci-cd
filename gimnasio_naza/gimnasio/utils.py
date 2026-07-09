"""
UTILIDADES PARA EXPORTACION DE REPORTES
Modulo con funciones para exportar datos a PDF y Excel
"""


from weasyprint import HTML, CSS
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.http import HttpResponse
from django.template.loader import render_to_string
from django.templatetags.static import static
from django.utils import timezone
from datetime import datetime
from django.conf import settings



# ====== EXPORTACION A PDF ======
def exportar_pdf(request, titulo, columnas, datos, nombre_archivo):  #  'título' -> 'titulo' (sin tilde)
    logo_url = request.build_absolute_uri(static('img/gym.jpeg'))  #  'logotipo_url' -> 'logo_url'

    # Crear contexto para el template
    contexto = {
        'titulo': titulo,
        'columnas': columnas,
        'datos': datos,
        'logo_url': logo_url,   #  ahora coincide con la variable definida arriba
        'now': timezone.now(),
    }

    html_string = render_to_string('reportes/reporte_pdf.html', contexto)

    html_object = HTML(string=html_string, base_url=request.build_absolute_uri('/'))  # base_url correcta

    pdf_bytes = html_object.write_pdf()

    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}.pdf"'

    return response

# ====== EXPORTACION A EXCEL ======
def exportar_excel(titulo, columnas, datos, nombre_archivo):
    """
    FUNCIÓN PARA EXPORTAR DATOS A EXCEL USANDO OPENPYXL
    
    Args:
        titulo: Titulo del reporte
        columnas: Lista de nombres de columnas
        datos: Lista de tuplas o diccionarios con los datos
        nombre_archivo: Nombre del archivo Excel a descargar
    
    Returns:
        HttpResponse con el archivo Excel generado
    """
    
    # Crear un nuevo libro de Excel
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Reporte"
    
    # Configurar estilos para el título
    title_font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
    title_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    title_alignment = Alignment(horizontal='center', vertical='center')
    
    # Agregar titulo
    worksheet.merge_cells('A1:' + chr(64 + len(columnas)) + '1')
    titulo_cell = worksheet['A1']
    titulo_cell.value = titulo
    titulo_cell.font = title_font
    titulo_cell.fill = title_fill
    titulo_cell.alignment = title_alignment
    worksheet.row_dimensions[1].height = 25
    
    # Configurar estilos para los encabezados
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Agregar encabezados de columnas
    for col_num, columna in enumerate(columnas, 1):
        cell = worksheet.cell(row=3, column=col_num)
        cell.value = columna
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    worksheet.row_dimensions[3].height = 20
    
    # Configurar estilos para los datos
    data_alignment = Alignment(horizontal='left', vertical='center')
    data_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Agregar datos al Excel
    data_fill_alternated = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    
    for row_num, fila in enumerate(datos, 4):
        # Convertir diccionario a tupla si es necesario
        if isinstance(fila, dict):
            valores = [fila.get(col.lower().replace(' ', '_'), '') for col in columnas]
        else:
            valores = fila
        
        # Llenar las celdas con datos
        for col_num, valor in enumerate(valores, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = valor
            cell.alignment = data_alignment
            cell.border = data_border
            
            # Colorear filas alternas para mejor legibilidad
            if (row_num - 4) % 2 == 0:
                cell.fill = data_fill_alternated
    
    # Ajustar ancho de columnas automaticamente
    for col_num, columna in enumerate(columnas, 1):
        max_length = len(str(columna))
        column_letter = chr(64 + col_num)
        
        for row in worksheet.iter_rows(min_col=col_num, max_col=col_num):
            for cell in row:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
        
        worksheet.column_dimensions[column_letter].width = max_length + 2
    
    # Crear respuesta HTTP con el Excel
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}.xlsx"'
    
    # Guardar el libro en la respuesta
    workbook.save(response)
    
    return response


# ====== GOOGLE FORMS API FUNCTIONS ======

import os
from django.conf import settings
from googleapiclient.discovery import build
from google.oauth2.credentials import Credentials
from google.auth.transport.requests import Request
from google_auth_oauthlib.flow import InstalledAppFlow
from gimnasio.googletoken import get_credentials

SCOPES = ['https://www.googleapis.com/auth/forms.body', 'https://www.googleapis.com/auth/forms.responses.readonly']


def get_google_forms_service():
    creds = get_credentials()  

    service = build('forms', 'v1', credentials=creds)
    return service


def validar_credenciales_google_forms():
    """
    Valida que el proyecto tenga credenciales válidas para Google Forms.
    Devuelve (True, mensaje) o (False, error)
    """

    credentials_path = os.path.join(settings.BASE_DIR, 'credentials.json')
    token_path = os.path.join(settings.BASE_DIR, 'token.json')

    # Verificar que exista credentials.json
    if not os.path.exists(credentials_path):
        return False, f"No se encuentra credentials.json en {settings.BASE_DIR}"

    # Verificar que exista token.json
    if not os.path.exists(token_path):
        return False, "No existe token.json. Debes generarlo una vez."

    try:
        creds = get_credentials()

        if creds and creds.valid:
            return True, "Credenciales válidas y activas."

        return False, "Credenciales inválidas."

    except Exception as e:
        return False, f"Error de autenticación: {str(e)}"

def crear_formulario_google(titulo, descripcion=""):
    """
    Crea un nuevo formulario en Google Forms.
    """
    service = get_google_forms_service()
    
    form = {
        "info": {
            "title": titulo,
            "description": descripcion,
        }
    }
    
    result = service.forms().create(body=form).execute()
    return result

def actualizar_formulario_google(form_id, updates):
    """
    Actualiza un formulario existente en Google Forms.
    updates debe ser un diccionario con los cambios.
    """
    service = get_google_forms_service()
    
    request = service.forms().batchUpdate(formId=form_id, body=updates)
    response = request.execute()
    return response

def obtener_formulario_google(form_id):
    """
    Obtiene la información de un formulario.
    """
    service = get_google_forms_service()
    
    result = service.forms().get(formId=form_id).execute()
    return result

def agregar_preguntas_a_formulario(form_id, formset):
    service = get_google_forms_service()

    form_actual = service.forms().get(formId=form_id).execute()
    index = len(form_actual.get("items", []))

    requests = []  #Lista para acumular todas las preguntas

    for form in formset:
        if form.cleaned_data and not form.cleaned_data.get('DELETE', False):

            pregunta = form.cleaned_data.get('pregunta')
            tipo = form.cleaned_data.get('tipo')
            opciones = form.cleaned_data.get('opciones', '')
            requerida = form.cleaned_data.get('requerida', False)

            if not pregunta:  
                continue

            if isinstance(opciones, str):
                opciones = [op.strip() for op in opciones.split(',') if op.strip()]

            question = {}

            if tipo == 'short_answer':
                question = {"textQuestion": {}}
            elif tipo == 'paragraph':
                question = {"textQuestion": {"paragraph": True}}
            elif tipo == 'multiple_choice':
                question = {
                    "choiceQuestion": {
                        "type": "RADIO",
                        "options": [{"value": op} for op in opciones]
                    }
                }
            elif tipo == 'check_boxes':
                question = {
                    "choiceQuestion": {
                        "type": "CHECKBOX",
                        "options": [{"value": op} for op in opciones]
                    }
                }
            elif tipo == 'dropdown':
                question = {
                    "choiceQuestion": {
                        "type": "DROP_DOWN",
                        "options": [{"value": op} for op in opciones]
                    }
                }
            elif tipo == 'linear_scale':
                question = {"scaleQuestion": {"low": 1, "high": 5}}
            elif tipo == 'date':
                question = {"dateQuestion": {}}
            elif tipo == 'time':
                question = {"timeQuestion": {}}

            requests.append({
                "createItem": {
                    "item": {
                        "title": pregunta,
                        "questionItem": {
                            "question": {
                                "required": requerida,
                                **question
                            }
                        }
                    },
                    "location": {"index": index}
                }
            })

            index += 1

    
    if requests:
        service.forms().batchUpdate(
            formId=form_id,
            body={"requests": requests}
        ).execute()