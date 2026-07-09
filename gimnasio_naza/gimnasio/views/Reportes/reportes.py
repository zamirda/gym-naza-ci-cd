from django.shortcuts import render
from django.views.generic import View
from django.views import View as DjangoView
from django.http import HttpResponse
from gimnasio.models import *
from gimnasio.utils import exportar_pdf, exportar_excel
from datetime import datetime
from gimnasio.views.registrovisitantestemporales.views import *
from gimnasio.views.turnosdeentrenadores.views import *
from gimnasio.views.certificacionesinternas.views import *
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from django.core.exceptions import ObjectDoesNotExist

class ExportarAsistenciaPDF(DjangoView):
    """
    VISTA PARA EXPORTAR ASISTENCIAS A PDF
    Obtiene todas las asistencias y las exporta en formato PDF
    """
    
    def get(self, request):
        # Obtener todas las asistencias 
        asistencia = Asistencia.objects.all()
        
        # Definir las columnas que se muestran en el reporte
        columnas = ['ID', 'Fecha de asistencia', 'Hora de ingreso', 'Membresia', 'Documento de usuario']
        
        # Preparar los datos en formato de tuplas (Formato de strings corregido)
        datos = [
            (
                a.id,
                a.fecha_asistencia,
                a.hora_ingreso,
                f"{a.fk_membresia.fk_usuario.nombre_usuario} {a.fk_membresia.fk_usuario.apellido_usuario}",
                a.fk_membresia.fk_usuario.documento
            )
            for a in asistencia
        ]
        
        # Generar nombre del archivo con timestamp
        nombre_archivo = f'Reporte_Asistencia_{datetime.now().strftime("%d_%m_%Y")}'
        
        # Controlar de forma segura si el usuario actual tiene perfil 'usuario'
        # --- BUSCA ESTA PARTE EN TU VISTA Y REEMPLÁZALA ---
        try:
            usuario_logueado = request.user.usuario
        except (ObjectDoesNotExist, AttributeError):
            # Opción rápida: Busca cualquier usuario de la tabla para que no sea null
            from gimnasio.models import Usuario  # Asegúrate de importar tu modelo Usuario
            usuario_logueado = Usuario.objects.first() 
            
            # Si no hay absolutamente ningún usuario creado en la tabla todavía:
            if not usuario_logueado:
                # Podrías crear un usuario fantasma o lanzar un mensaje amigable
                pass

        Reportes_estadisticas.objects.create(
            tipo_reporte="asistencia",
            descripcion="Se generó el reporte de asistencias en PDF",
            fk_usuario=usuario_logueado  # Ya nunca será None si hay al menos un perfil
        )
        
        # Llamar funcion de exportacion a PDF
        return exportar_pdf(
            request,
            titulo='REPORTE DE ASISTENCIAS',
            columnas=columnas,
            datos=datos,
            nombre_archivo=nombre_archivo
        )


class ExportarAsistenciaExcel(DjangoView):
    """
    VISTA PARA EXPORTAR ASISTENCIAS A EXCEL
    Obtiene todas las asistencias y las exporta en formato Excel
    """
    
    def get(self, request):
        # Obtener todas las asistencias 
        asistencia = Asistencia.objects.all()
        
        # Definir las columnas que se muestran en el reporte
        columnas = ['ID', 'Fecha de asistencia', 'Hora de ingreso', 'Membresia', 'Documento de usuario']
        
        # Preparar los datos en formato de tuplas
        datos = [
            (
                a.id,
                a.fecha_asistencia,
                a.hora_ingreso,
                f"{a.fk_membresia.fk_usuario.nombre_usuario} {a.fk_membresia.fk_usuario.apellido_usuario}",
                a.fk_membresia.fk_usuario.documento
            )
            for a in asistencia
        ]
        
        # Generar nombre del archivo con timestamp
        nombre_archivo = f'Reporte_Asistencia_{datetime.now().strftime("%d_%m_%Y")}'
        
        # --- SOLUCIÓN DE RESPALDO PARA EVITAR INTEGRITYERROR ---
        try:
            usuario_logueado = request.user.usuario
        except (ObjectDoesNotExist, AttributeError):
            # Como 'juan' no tiene perfil asignado, importamos tu modelo de perfiles
            from gimnasio.models import Usuario  
            # Buscamos el primer perfil real del sistema para usarlo de puente
            usuario_logueado = Usuario.objects.first()
            
        # Si la tabla Usuario está completamente vacía en tu base de datos, 
        # esto podría seguir enviando None. Asegúrate de tener al menos un registro en la tabla Usuario.
        
        Reportes_estadisticas.objects.create(
            tipo_reporte="asistencia",
            descripcion="Se generó el reporte de asistencias en Excel",
            fk_usuario=usuario_logueado
        )
        
        # Llamar funcion de exportacion a Excel
        return exportar_excel(
            titulo='REPORTE DE ASISTENCIAS',
            columnas=columnas,
            datos=datos,
            nombre_archivo=nombre_archivo
        )
# ====== VISTAS PARA EXPORTAR REPORTES MEMBRESIA======

class ExportarMembresiaPDF(DjangoView):
    """
    VISTA PARA EXPORTAR CATEGORIAS A PDF
    Obtiene todas las categorías y las exporta en formato PDF
    """
    
    def get(self, request):
        # Obtener todas las categorias 
        membresia = Membresia.objects.all()
        
        # Definir las columnas que se 
        # 
        # en el reporte
        columnas = ['ID','Fecha de inicio','Fecha de fin','Membresia','Documento de usuario']
        
        # Preparar los datos en formato de tuplas
        datos = [
            (m.id,m.fecha_inicio, m.fecha_fin,f"{m.fk_usuario.nombre_usuario,m.fk_usuario.apellido_usuario}",m.fk_usuario.documento)
            for m in membresia
        ]
        
        # Generar nombre del archivo con timestamp
        nombre_archivo = f'Reporte_membresia_{datetime.now().strftime("%d_%m_%Y")}'
        
        try:
            usuario_logueado = request.user.usuario
        except (ObjectDoesNotExist, AttributeError):
            # Opción rápida: Busca cualquier usuario de la tabla para que no sea null
            from gimnasio.models import Usuario  # Asegúrate de importar tu modelo Usuario
            usuario_logueado = Usuario.objects.first() 
            
            # Si no hay absolutamente ningún usuario creado en la tabla todavía:
            if not usuario_logueado:
                # Podrías crear un usuario fantasma o lanzar un mensaje amigable
                pass
            
            
        Reportes_estadisticas.objects.create(
            tipo_reporte="membresia",
            tipo_archivo="PDF",
            descripcion="Se generó el reporte de membresías en PDF",
            fk_usuario=request.user.usuario
        )
        # Llamar funcion de exportacion a PDF
        return exportar_pdf(
            request,
            titulo='REPORTE DE MEMBRESIAS',
            columnas=columnas,
            datos=datos,
            nombre_archivo=nombre_archivo
        )


class ExportarMembresiaExcel(DjangoView):
    """
    VISTA PARA EXPORTAR CATEGORIAS A EXCEL
    Obtiene todas las categorias y las exporta en formato Excel
    """
    
    def get(self, request):
        # Obtener todas las categorias 
        membresia = Membresia.objects.all()
        
        # Definir las columnas que se 
        # 
        # en el reporte
        columnas = ['ID','Fecha de inicio','Fecha de fin','Membresia','Documento de usuario']
        
        # Preparar los datos en formato de tuplas
        datos = [
            (m.id,m.fecha_inicio, m.fecha_fin,f"{m.fk_usuario.nombre_usuario,m.fk_usuario.apellido_usuario}",m.fk_usuario.documento)
            for m in membresia
        ]
        
        # Generar nombre del archivo con timestamp
        nombre_archivo = f'Reporte_membresia_{datetime.now().strftime("%d_%m_%Y")}'    
        Reportes_estadisticas.objects.create(
            tipo_reporte="membresia",
            tipo_archivo="EXCEL",
            descripcion="Se generó el reporte de membresías en Excel",
            fk_usuario=request.user.usuario
        )
        # Llamar funcion de exportacion a PDF
        return exportar_excel(
            titulo='REPORTE DE MEMBRESIAS',
            columnas=columnas,
            datos=datos,
            nombre_archivo=nombre_archivo
        )


class ExportarElementosPDF(DjangoView):
    """
    VISTA PARA EXPORTAR ELEMENTOS A PDF
    Obtiene todos los elementos y los exporta en formato PDF
    """
    
    def get(self, request):
        # Obtener todas las categorias 
        elemento= Elemento.objects.all()
        
        # Definir las columnas que se mostraran en el reporte
        columnas = ['ID', 'Nombre','marca','categoria']
        
        # Preparar los datos en  tuplas
        datos = [
            (e.id, e.nombre_elemento,e.marca,e.nombre_categoria.nombre_categoria)
            for e in elemento
        ]
        
        # Generar nombre del archivo con timestamp
        nombre_archivo = f'Reporte_Elementos_{datetime.now().strftime("%d_%m_%Y")}'

        Reportes_estadisticas.objects.create(
            tipo_reporte="elemento",
            tipo_archivo="PDF",
            descripcion="Se generó el reporte de elementos en PDF",
            fk_usuario=request.user.usuario
        )
        # Llamar funcion de exportacion a PDF
        return exportar_pdf(
            request,
            titulo='REPORTE DE ELEMENTOS',
            columnas=columnas,
            datos=datos,
            nombre_archivo=nombre_archivo
        )
class ExportarElementosExcel(DjangoView):
    """
    VISTA PARA EXPORTAR ELEMENTOS A EXCEL
    Obtiene todos los elementos y los exporta en formato Excel
    """
    
    def get(self, request):
        # Obtener todas las categorias 
        elemento= Elemento.objects.all()
        
        # Definir las columnas que se mostraran en el reporte
        columnas = ['ID', 'Nombre', 'Marca', 'Categoria']
        
        # Preparar los datos en  tuplas
        datos = [
            (e.id, e.nombre_elemento, e.marca , e.nombre_categoria.nombre_categoria)
            for e in elemento
        ]
        
        # Generar nombre del archivo con timestamp
        nombre_archivo = f'Reporte_Elementos_{datetime.now().strftime("%d_%m_%Y")}'
        Reportes_estadisticas.objects.create(
            tipo_reporte="elemento",
            tipo_archivo="EXCEL",
            descripcion="Se generó el reporte de elementos en Excel",
            fk_usuario=request.user.usuario
        )
        # Llamar funcion de exportacion a Excel
        return exportar_excel(
            titulo='REPORTE DE ELEMENTOS',
            columnas=columnas,
            datos=datos,
            nombre_archivo=nombre_archivo
        )
class ExportarUsuariosPDF(DjangoView):
    """
    VISTA PARA EXPORTAR USUARIOS A PDF
    Obtiene todos los usuarios y los exporta en formato PDF
    """
    
    def get(self, request):
        # Obtener todos los usuarios
        usuarios = Usuario.objects.all()
        
        # Definir las columnas que se mostraran en el reporte
        columnas = ['ID', 'documento','genero','nombre ','apellido','correo de','telefono','fecha de nacimiento','estado']
        
        # Preparar los datos en  tuplas
        datos = [
            (u.id, u.documento, u.genero_usuario, u.nombre_usuario, u.apellido_usuario, u.correo_usuario, u.telefono_usuario, u.fecha_nacimiento, u.estado)
            for u in usuarios
        ]
        
        # Generar nombre del archivo con timestamp
        nombre_archivo = f'Reporte_Usuarios_{datetime.now().strftime("%d_%m_%Y")}'
        Reportes_estadisticas.objects.create(
            tipo_reporte="usuarios",
            tipo_archivo="PDF",
            descripcion="Se generó el reporte de usuarios en PDF",
            fk_usuario=request.user.usuario
        )
        # Llamar funcion de exportacion a PDF
        return exportar_pdf(
            request,
            titulo='REPORTE DE USUARIOS',
            columnas=columnas,
            datos=datos,
            nombre_archivo=nombre_archivo
        ) 
class ExportarUsuariosExcel(DjangoView):
    """
    VISTA PARA EXPORTAR USUARIOS A EXCEL
    Obtiene todos los usuarios y los exporta en formato Excel
    """
    
    def get(self, request):
        # Obtener todos los usuarios
        usuarios = Usuario.objects.all()
        
        # Definir las columnas que se mostraran en el reporte
        columnas = ['ID', 'documento','genero del usuario','nombre del usuario','apellido del usuario','correo del usuario','telefono del usuario','fecha de nacimiento del usuario','estado del usuario']
        
        # Preparar los datos en  tuplas
        datos = [
            (u.id, u.documento, u.genero_usuario, u.nombre_usuario, u.apellido_usuario, u.correo_usuario, u.telefono_usuario, u.fecha_nacimiento, u.estado)
            for u in usuarios
        ]
        
        # Generar nombre del archivo con timestamp
        nombre_archivo = f'Reporte_Usuarios_{datetime.now().strftime("%d_%m_%Y")}'
        Reportes_estadisticas.objects.create(
            tipo_reporte="usuarios",
            tipo_archivo="EXCEL",
            descripcion="Se generó el reporte de usuarios en Excel",
            fk_usuario=request.user.usuario
        )
        # Llamar funcion de exportacion a Excel
        return exportar_excel(
            titulo='REPORTE DE USUARIOS',
            columnas=columnas,
            datos=datos,
            nombre_archivo=nombre_archivo
        )
class ExportarmantenimientoPDF(DjangoView):
    """
    VISTA PARA EXPORTAR MANTENIMIENTO A PDF
    Obtiene todos los mantenimientos y los exporta en formato PDF
    """
    
    def get(self, request):
        # Obtener todos los mantenimientos
        mantenimientos = Mantenimiento.objects.all()
        
        # Definir las columnas que se mostraran en el reporte
        columnas = ['ID', 'nombre_elemento', 'Tipo de mantenimiento', 'Fecha programada', 'Estado']
        
        # Preparar los datos en  tuplas
        datos = [
            (m.id, m.nombre_elemento, m.tipo_mantenimiento, m.fecha_programada, m.estado)
            for m in mantenimientos
        ]
        
        # Generar nombre del archivo con timestamp
        nombre_archivo = f'Reporte_Mantenimiento_{datetime.now().strftime("%d_%m_%Y")}'
        Reportes_estadisticas.objects.create(
            tipo_reporte="mantenimiento",
            tipo_archivo="PDF",
            descripcion="Se generó el reporte de mantenimientos en PDF",
            fk_usuario=request.user.usuario
        )
        # Llamar funcion de exportacion a PDF
        return exportar_pdf(
            request,
            titulo='REPORTE DE MANTENIMIENTO',
            columnas=columnas,
            datos=datos,
            nombre_archivo=nombre_archivo
        )
class ExportarMantenimientoExcel(DjangoView):
    """
    VISTA PARA EXPORTAR MANTENIMIENTO A EXCEL
    Obtiene todos los mantenimientos y los exporta en formato Excel
    """
    
    def get(self, request):
        # Obtener todos los mantenimientos
        mantenimientos = Mantenimiento.objects.all()
        
        # Definir las columnas que se mostraran en el reporte
        columnas = ['ID', 'nombre elemento', 'Tipo de mantenimiento', 'Fecha programada', 'Estado']
        
        # Preparar los datos en  tuplas
        datos = [
            (m.id, str(m.nombre_elemento), m.tipo_mantenimiento, m.fecha_programada, m.estado)
            for m in mantenimientos
            
        ]
        
        # Generar nombre del archivo con timestamp
        nombre_archivo = f'Reporte_Mantenimiento_{datetime.now().strftime("%d_%m_%Y")}'
        Reportes_estadisticas.objects.create(
            tipo_reporte="mantenimiento",
            tipo_archivo="EXCEL",
            descripcion="Se generó el reporte de mantenimientos en Excel",
            fk_usuario=request.user.usuario
        )
        # Llamar funcion de exportacion a Excel
        return exportar_excel(
            titulo='REPORTE DE MANTENIMIENTO',
            columnas=columnas,
            datos=datos,
            nombre_archivo=nombre_archivo
        )
class ExportarRegistrovisitantestemporalesPDF(DjangoView):
    """
    VISTA PARA EXPORTAR CERTIFICACIONES INTERNAS A PDF
    Obtiene todas las certificaciones internas y las exporta en formato PDF
    """
    
    def get(self, request):
        # Obtener todas las certificaciones internas
        registro = Registrovisitantestemporales.objects.all()
        
        # Definir las columnas que se muestran en el reporte
        columnas = ['ID', 'fecha_registro' ]
        
        # Preparar los datos en formato de tuplas
        datos = [
            (visitante.id, visitante.fecha_registro, )
            for visitante in registro
        ]
        
        # Generar nombre del archivo con timestamp
        nombre_archivo = f'Registrovisitantestemporales{datetime.now().strftime("%d_%m_%Y")}'
        Reportes_estadisticas.objects.create(
            tipo_reporte="visitantes",
            tipo_archivo="PDF",
            descripcion="Se generó el reporte de visitantes temporales en PDF",
            fk_usuario=request.user.usuario
        )
        # Llamar funcion de exportacion a PDF
        return exportar_pdf(
            request,
            titulo='REPORTE DE REGISTRO DE VISITANTES TEMPORALES',
            columnas=columnas,
            datos=datos,
            nombre_archivo=nombre_archivo
        )


class ExportarRegistrovisitantestemporalesExcel(DjangoView):
    """
    VISTA PARA EXPORTAR CATEGORIAS A EXCEL
    Obtiene todas las categorias y las exporta en formato Excel
    """
    
    def get(self, request):
        # Obtener todas las certificaciones internas
        registro = Registrovisitantestemporales.objects.all()
        
        # Definir las columnas que se mostraran en el reporte
        columnas = ['ID', 'fecha_registro']
        
        # Preparar los datos en  tuplas
        datos = [
            (visitante.id, visitante.fecha_registro, )
            for visitante in registro
        ]
        
        # Generar nombre del archivo con timestamp
        nombre_archivo = f'Registrovisitantestemporales{datetime.now().strftime("%d_%m_%Y")}'
        Reportes_estadisticas.objects.create(
            tipo_reporte="visitantes",
            tipo_archivo="EXCEL",
            descripcion="Se generó el reporte de visitantes temporales en Excel",
            fk_usuario=request.user.usuario
        )
        # Llamar funcion de exportacion a Excel
        return exportar_excel(
            titulo='REPORTE DE REGISTRO DE VISITANTES TEMPORALES',
            columnas=columnas,
            datos=datos,
            nombre_archivo=nombre_archivo
        )

class ExportarTurnodeentrenadorPDF(DjangoView):
    """
    VISTA PARA EXPORTAR TURNO DE ENTRENADOR A PDF
    Obtiene todos los turnos de entrenador y los exporta en formato PDF
    """
    
    def get(self, request):
        # Obtener todas las certificaciones internas
        registro = Turnosentrenadores.objects.all()
        
        # Definir las columnas que se muestran en el reporte
        columnas = ['ID', 'Administrador', 'fecha_turno_inicio', 'fecha_turno_final', 'Jornada']
        
        # Preparar los datos en formato de tuplas
        datos = [
            (turno.id, turno.administrador.nombre_usuario if turno.administrador else "Sin administrador", turno.fecha_turno_inicio, turno.fecha_turno_final, turno.jornada)
            for turno in registro
        ]
        
        # Generar nombre del archivo con timestamp
        nombre_archivo = f'Turnodeentrenador {datetime.now().strftime("%d_%m_%Y")}'
        Reportes_estadisticas.objects.create(
            tipo_reporte="turnos",
            tipo_archivo="PDF",
            descripcion="Se generó el reporte de turnos de entrenadores en PDF",
            fk_usuario=request.user.usuario
        )
        # Llamar funcion de exportacion a PDF
        return exportar_pdf(
            request,
            titulo='REPORTE DE TURNO DE ENTRENADOR ',
            columnas=columnas,
            datos=datos,
            nombre_archivo=nombre_archivo
        )
        
class ExportarTurnodeentrenadorExcel(DjangoView):
    """
    VISTA PARA EXPORTAR CATEGORIAS A EXCEL
    Obtiene todas las categorias y las exporta en formato Excel
    """
    
    def get(self, request):
        # Obtener todas las certificaciones internas
        registro = Turnosentrenadores.objects.all()
        
        # Definir las columnas que se mostraran en el reporte
        columnas = ['ID', 'Administrador', 'fecha_turno_inicio', 'fecha_turno_final', 'Jornada']
        
        # Preparar los datos en  tuplas
        datos = [
            (turno.id, turno.administrador.nombre_usuario if turno.administrador else "Sin administrador", turno.fecha_turno_inicio, turno.fecha_turno_final, turno.jornada)
            for turno in registro
        ]
        
        # Generar nombre del archivo con timestamp
        nombre_archivo = f'Turnodeentrenador{datetime.now().strftime("%d_%m_%Y")}'
        Reportes_estadisticas.objects.create(
            tipo_reporte="turnos",
            tipo_archivo="EXCEL",
            descripcion="Se generó el reporte de turnos de entrenadores en Excel",
            fk_usuario=request.user.usuario
        )
        # Llamar funcion de exportacion a Excel
        return exportar_excel(
            titulo='REPORTE DE TURNO DE ENTRENADOR',
            columnas=columnas,
            datos=datos,
            nombre_archivo=nombre_archivo
        )


class ExportarCertificacioninternaPDF(DjangoView):
    """
    VISTA PARA EXPORTAR CERTIFICACIONES INTERNAS A PDF
    Obtiene todas las certificaciones internas y las exporta en formato PDF
    """
    
    def get(self, request):
        # Obtener todas las certificaciones internas
        registro = Certificacion_interna.objects.all()
        
        # Definir las columnas que se muestran en el reporte
        columnas = ['ID', 'descripcion_certificacion', 'fecha_certificacion', 'fk_Asistencia']
        
        # Preparar los datos en formato de tuplas
        datos = [
    (
        turno.id,
        turno.descripcion_certificacion,
        turno.fecha_certificacion,
        str(turno.fk_membresia)
    )
    for turno in registro
]
        
        # Generar nombre del archivo con timestamp
        nombre_archivo = f'CertificacionesInternas {datetime.now().strftime("%d_%m_%Y")}'
        Reportes_estadisticas.objects.create(
            tipo_reporte="certificaciones",
            tipo_archivo="PDF",
            descripcion="Se generó el reporte de certificaciones en PDF",
            fk_usuario=request.user.usuario
        )
        # Llamar funcion de exportacion a PDF
        return exportar_pdf(
            request,
            titulo='REPORTE DE CERTIFICACIONES INTERNAS',
            columnas=columnas,
            datos=datos,
            nombre_archivo=nombre_archivo
        )
        
class ExportarCertificacioninternaExcel(DjangoView):
    """
    VISTA PARA EXPORTAR CATEGORIAS A EXCEL
    Obtiene todas las categorias y las exporta en formato Excel
    """
    
    def get(self, request):
        registro = Certificacion_interna.objects.all()

        wb = Workbook()
        ws = wb.active
        ws.title = "Certificaciones Internas"
        # Obtener todas las certificaciones internas
        registro = Certificacion_interna.objects.all()
        
        # Definir las columnas que se mostraran en el reporte
        columnas = ['ID', 'descripcion_certificacion', 'fecha_certificacion', 'fk_Asistencia']
        
        # Preparar los datos en  tuplas
        datos = [
            (turno.id, turno.descripcion_certificacion, turno.fecha_certificacion, str(turno.fk_membresia))
            for turno in registro
        ]
        
        # Generar nombre del archivo con timestamp
        nombre_archivo = f'CertificacionesInternas {datetime.now().strftime("%d_%m_%Y")}'
        Reportes_estadisticas.objects.create(
            tipo_reporte="certificaciones",
            tipo_archivo="EXCEL",
            descripcion="Se generó el reporte de certificaciones en Excel",
            fk_usuario=request.user.usuario
        )
        # Llamar funcion de exportacion a Excel
        return exportar_excel(
            titulo='REPORTE DE CERTIFICACIONES INTERNAS',
            columnas=columnas,
            datos=datos,
            nombre_archivo=nombre_archivo
        )
        
        
class ExportarReportes_estadisticasPDF(DjangoView):
    """
    VISTA PARA EXPORTAR CATEGORIAS A PDF
    Obtiene todas las categorías y las exporta en formato PDF
    """
    
    def get(self, request):
        # Obtener todas las categorias 
        reportes = Reportes_estadisticas.objects.all()
        
        # Definir las columnas que se muestran en el reporte
        columnas = ['ID', 'Tipo Reporte', 'Fecha Generacion']
        
        # Preparar los datos en formato de tuplas
        datos = [
            (rep.id, rep.tipo_reporte, rep.fecha_generacion)
            for rep in reportes
        ]
        
        # Generar nombre del archivo con timestamp
        nombre_archivo = f'Reportes_Estadisticas{datetime.now().strftime("%d_%m_%Y")}'
        
        # Llamar funcion de exportacion a PDF
        return exportar_pdf(
            request,
            titulo='REPORTE Y ESTADISTICAS',
            columnas=columnas,
            datos=datos,
            nombre_archivo=nombre_archivo
        )


class ExportarReportes_estadisticasEXCEL(DjangoView):
    """
    VISTA PARA EXPORTAR CATEGORIAS A EXCEL
    Obtiene todas las categorias y las exporta en formato Excel
    """
    
    def get(self, request):
        # Obtener todas las categorias 
        reportes = Reportes_estadisticas.objects.all()
        
        # Definir las columnas que se mostraran en el reporte
        columnas = ['ID', 'Tipo Reporte', 'Fecha Generacion']
        
        # Preparar los datos en  tuplas
        datos = [
            (rep.id, rep.tipo_reporte, rep.fecha_generacion)
            for rep in reportes
        ]
        
        # Generar nombre del archivo con timestamp
        nombre_archivo = f'Reportes_estadisticas{datetime.now().strftime("%d_%m_%Y")}'
        
        # Llamar funcion de exportacion a Excel
        return exportar_excel(
            titulo='REPORTE Y ESTADISTICAS',
            columnas=columnas,
            datos=datos,
            nombre_archivo=nombre_archivo
        )

# ====== VISTAS PARA EXPORTAR SOPORTES Y PQRS ======

class ExportarSoporte_PQRSPDF(DjangoView):
    """
    VISTA PARA EXPORTAR CATEGORIAS A PDF
    Obtiene todas las categorías y las exporta en formato PDF
    """
    
    def get(self, request):
        # Obtener todas las categorias 
        soportes = Soporte_PQRS.objects.all()
        
        # Definir las columnas que se muestran en el reporte
        columnas = ['ID','Tipo', 'Descripcion', 'Fecha Ingreso', 'Estado']
        
        # Preparar los datos en formato de tuplas
        datos = [
            (soporte.id, soporte.tipo, soporte.descripcion, soporte.fecha_ingreso, soporte.estado)
            for soporte in soportes
        ]
        
        # Generar nombre del archivo con timestamp
        nombre_archivo = f'Soporte_PQRS{datetime.now().strftime("%d_%m_%Y")}'
        
        # Llamar funcion de exportacion a PDF
        return exportar_pdf(
            request,
            titulo='SOPORTE Y PQRS',
            columnas=columnas,
            datos=datos,
            nombre_archivo=nombre_archivo
        )


class ExportarSoporte_PQRSEXCEL(DjangoView):
    """
    VISTA PARA EXPORTAR CATEGORIAS A EXCEL
    Obtiene todas las categorias y las exporta en formato Excel
    """
    
    def get(self, request):
        # Obtener todas las categorias 
        soportes = Soporte_PQRS.objects.all()
        
        # Definir las columnas que se muestran en el soporte y pqr
        columnas = ['ID','Tipo', 'Descripcion', 'Fecha Ingreso', 'Estado']
        
        # Preparar los datos en formato de tuplas
        datos = [
            (soporte.id, soporte.tipo, soporte.descripcion, soporte.fecha_ingreso, soporte.estado)
            for soporte in soportes
        ]
        
        # Generar nombre del archivo con timestamp
        nombre_archivo = f'Soporte_PQRS{datetime.now().strftime("%d_%m_%Y")}'
        
        # Llamar funcion de exportacion a Excel
        return exportar_excel(
            titulo='SOPORTE Y PQRS',
            columnas=columnas,
            datos=datos,
            nombre_archivo=nombre_archivo
        )
class ExportarSancionPDF(DjangoView):
    
    def get(self, request):
        sanciones = Sancion.objects.all()

        columnas = ['Fecha', 'Nombre', 'Cédula', 'Motivo de sanción']

        datos = [
            (
                s.fecha_inicio,
                s.fk_usuario.nombre_usuario,
                s.fk_usuario.documento,
                s.motivo_sancion
            )
            for s in sanciones
        ]

        nombre_archivo = f'Reporte_sanciones_{datetime.now().strftime("%d_%m_%Y")}'

        return exportar_pdf(
            request,
            titulo='REPORTE DE SANCIONES',
            columnas=columnas,
            datos=datos,
            nombre_archivo=nombre_archivo
        )
class ExportarSancionExcel(DjangoView):
    
    def get(self, request):

        sanciones = Sancion.objects.all()

        wb = Workbook()
        ws = wb.active
        ws.title = "Sanciones"

        # Encabezados
        columnas = ['Fecha', 'Nombre', 'Cédula', 'Motivo de sanción']
        ws.append(columnas)

        # Estilo encabezados
        for col in ws[1]:
            col.font = Font(bold=True, color="FFFFFF")
            col.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            col.alignment = Alignment(horizontal="center")

        # Datos
        for s in sanciones:
            ws.append([
                s.fecha_inicio.strftime("%d/%m/%Y"),
                s.fk_usuario.nombre_usuario,
                s.fk_usuario.documento,
                s.motivo_sancion
            ])

        # Ajustar ancho columnas automáticamente
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            ws.column_dimensions[column_letter].width = max_length + 2

        nombre_archivo = f'Reporte_sanciones_{datetime.now().strftime("%d_%m_%Y")}.xlsx'

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        response['Content-Disposition'] = f'attachment; filename={nombre_archivo}'

        wb.save(response)

        return response